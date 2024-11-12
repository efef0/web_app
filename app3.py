#23456----------------------------------------------------------------72
import streamlit as st
from openpyxl import load_workbook
from io import BytesIO
import os
from PIL import Image as PILImage
from openpyxl.drawing.image import Image
#23456----------------------------------------------------------------72
agree1 = st.checkbox("データを入力しますか？")
if agree1:
#23456----------------------------------------------------------------72
    wb = load_workbook('./data/mesure.xlsx')
    ws = wb.active
#23456----------------------------------------------------------------72
    pname = st.text_input('記録者氏名を入力してください')
    mname = st.selectbox('記録する機器名を選択してください',('No.1ミル','No.2ミル','No.3ミル','No.4ミル'))
    date = st.date_input('日付を選択してください')
    level1 = st.number_input('油面レベル(mm)を入力してください',value=0,step=1)
#23456----------------------------------------------------------------72
    agree2 = st.checkbox("その他異常はありましたか？")
    if agree2:
        problem = st.text_area('異常の内容を入力してください')
        ws['B8'].value = problem
    else :
        ws['B8'].value = '特に異常な状態は見られなかった'

    if st.button('Excel Fileにデータを転記する'):
        ws['G1'].value = pname
        ws['C3'].value = date
        ws['C4'].value = mname
        ws['C5'].value = level1

        filename = f'./data/{pname}.xlsx'
        wb.save(filename)
#23456----------------------------------------------------------------72
agree4 = st.checkbox("Excel Fileに写真を添付しますか？")
if agree4:
    pname3 = st.text_input('Excel File作成者の氏名を入力してください')
    filename4 = f'./data/{pname3}.xlsx'
    dl = st.checkbox("File作成者の氏名を入力しましたか？")
    if dl:
        wb3 = load_workbook(filename4)
        ws3 = wb3.active
        picture = st.camera_input("写真を撮影する")
        agree5 = st.checkbox("写真を保存しますか？")
        if agree5:
            temp_dir = "temp_images"
            os.makedirs(temp_dir, exist_ok=True)
            temp_image_path = os.path.join(temp_dir, "temp_image_.png")

            pil_img = PILImage.open(picture)
            pil_img = pil_img.resize((int(pil_img.width * 0.5), int(pil_img.height * 0.5)))
            pil_img.save(temp_image_path)
            img = Image(temp_image_path)

            ws3.add_image(img, "B13")

            wb3.save(filename4)

            os.remove(temp_image_path)
            os.rmdir(temp_dir)
#23456----------------------------------------------------------------72
agree3 = st.checkbox("Excel Fileをダウンロードしますか？")
if agree3:
    pname2 = st.text_input('Excel File作成者の氏名を入力してください')
    dl = st.checkbox("File作成者の氏名を入力しましたか？")
    if dl:
        filename2 = f'./data/{pname2}.xlsx'
        filename3 = f'{pname2}.xlsx'
        wb2 = load_workbook(filename2)
        output = BytesIO()
        wb2.save(output)
        excel_data2 = output.getvalue()
        output.close()
        st.download_button(label="Download", data=excel_data2, file_name=filename3)
        agree4 = st.checkbox("Excel Fileのダウンロードが完了したら、チェックしてください(当該ファイルをサーバーから消去します)")
        if agree4:
            os.remove(filename2)
